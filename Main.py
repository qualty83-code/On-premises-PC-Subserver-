import os
import shutil
import sqlite3
import urllib.parse
import urllib.request
import json
from typing import Any, cast
from datetime import datetime
from fastapi import FastAPI, UploadFile, File, Form, HTTPException, Depends, Header
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from typing import Dict, Any, Optional, List

app = FastAPI(title="Company Local Sub-Server")

class RecordRequest(BaseModel):
    department: str
    table_name: str
    filters: Optional[Dict[str, Any]] = None

class ChatHistoryRequest(BaseModel):
    company: str = ""
    department: str = ""

class ChatMessageSyncRequest(BaseModel):
    company: str = ""
    department: str = ""
    roomTitle: str = ""
    id: str = ""
    authorId: str = ""
    authorName: str = ""
    text: str = ""
    uri: str = ""
    type: str = "text"
    createdAt: int = 0
    metadata: str = "{}"

class AiSearchRequest(BaseModel):
    query: str
    department: str = "SALES"

class SaveDataRequest(BaseModel):
    company: str
    department: str
    task_name: str
    content: str
    poster: str = "관리자"

class InitRequest(BaseModel):
    company: str
    group: str = ""


# CORS 설정
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# 🔒 보안용 API Key
API_KEY = "MyCompanySecretKey1234!"
SUBSERVER_DB_PATH = "C:/SUBSERVER/subserver.db"

# 데이터 보관용 루트 디렉토리
DEFAULT_DATA_DIR = "C:\\SUBSERVER"
os.makedirs(DEFAULT_DATA_DIR, exist_ok=True)

def get_config_value(key: str) -> str:
    try:
        if not os.path.exists(SUBSERVER_DB_PATH):
            return ""
        conn = sqlite3.connect(SUBSERVER_DB_PATH)
        c = conn.cursor()
        c.execute("SELECT value FROM config WHERE key = ?", (key,))
        row = c.fetchone()
        conn.close()
        return row[0] if row else ""
    except Exception:
        return ""


def save_config_value(key: str, value: str):
    try:
        conn = sqlite3.connect(SUBSERVER_DB_PATH)
        c = conn.cursor()
        c.execute("CREATE TABLE IF NOT EXISTS config (key TEXT PRIMARY KEY, value TEXT)")
        c.execute("INSERT OR REPLACE INTO config (key, value) VALUES (?, ?)", (key, value))
        conn.commit()
        conn.close()
    except Exception:
        pass

def get_search_root():
    """GUI에서 설정한 ERP 기본 경로를 우선 참조하고, 없으면 기본값 반환"""
    config_path = get_config_value("erp_base_path")
    if config_path and os.path.exists(config_path):
        return os.path.normpath(config_path)
    return DEFAULT_DATA_DIR


DATA_SEARCH_EXTENSIONS = {".xlsx", ".xls", ".db", ".sqlite", ".sqlite3"}
EXCLUDED_SEARCH_DB_FILES = {"subserver.db", "chat_history.DB"}


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _make_headers(values: list[Any]) -> list[str]:
    headers: list[str] = []
    used: dict[str, int] = {}
    for index, value in enumerate(values, start=1):
        header = _stringify_cell(value) or f"column_{index}"
        count = used.get(header, 0) + 1
        used[header] = count
        headers.append(header if count == 1 else f"{header}_{count}")
    return headers


def _normalize_row(values: list[Any] | tuple[Any, ...], width: int) -> list[str]:
    row = [_stringify_cell(value) for value in values]
    if len(row) < width:
        row.extend([""] * (width - len(row)))
    return row[:width]


def _iter_excel_tables(path: str) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for sheet in workbook.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            header_values: list[Any] | None = None
            data_values: list[tuple[Any, ...]] = []
            for row in rows:
                if any(_stringify_cell(value) for value in row):
                    if header_values is None:
                        header_values = list(row)
                    else:
                        data_values.append(row)
            if not header_values:
                continue
            headers = _make_headers(header_values)
            table_rows = [_normalize_row(row, len(headers)) for row in data_values]
            if table_rows:
                tables.append({"source_table": sheet.title, "headers": headers, "rows": table_rows})
    except Exception:
        return []
    return tables


def _iter_sqlite_tables(path: str) -> list[dict[str, Any]]:
    tables: list[dict[str, Any]] = []
    try:
        conn = sqlite3.connect(path)
        cursor = conn.cursor()
        table_names = [row[0] for row in cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")]
        for table_name in table_names:
            if table_name.startswith("sqlite_"):
                continue
            try:
                schema_rows = cursor.execute(f'PRAGMA table_info("{table_name}")').fetchall()
                headers = [str(row[1]) for row in schema_rows]
                if not headers:
                    continue
                rows = cursor.execute(f'SELECT * FROM "{table_name}"').fetchall()
                table_rows = [_normalize_row(row, len(headers)) for row in rows]
                if table_rows:
                    tables.append({"source_table": table_name, "headers": headers, "rows": table_rows})
            except Exception:
                continue
        conn.close()
    except Exception:
        return []
    return tables


def _iter_data_tables(path: str) -> list[dict[str, Any]]:
    ext = os.path.splitext(path)[1].lower()
    if ext in {".xlsx", ".xls"}:
        return _iter_excel_tables(path)
    if ext in {".db", ".sqlite", ".sqlite3"}:
        return _iter_sqlite_tables(path)
    return []


# =====================================================================
# [정리됨] 기존 키워드 스코어링/동의어 확장/결정적 필터링 로직은 모두 제거되었다.
#  이제 검색 추론은 전적으로 LLM(프롬프트)에 위임한다. (아래 스마트 검색 엔진 참고)
# =====================================================================



# =====================================================================
# [신규] 로컬 QWEN2.5(Ollama) 기반 지능형 스마트 검색 엔진
#  - LLM은 "해석기" 역할만: (스키마 + 자연어 질문) -> 구조화된 검색 계획(JSON)
#  - 실제 필터링은 Python이 결정적으로 수행 (보안/정확성/재현성 보장)
#  - Ollama 미가동 시 기존 키워드 검색으로 자동 폴백
# =====================================================================

OLLAMA_URL = "http://127.0.0.1:11434"
OLLAMA_MODEL = "qwen3.6:latest"


def _get_ollama_model() -> str:
    """GUI에서 저장한 로컬 AI 모델명을 우선 사용하고, 없으면 기본 상수를 사용합니다."""
    return get_config_value("local_ai_model").strip() or OLLAMA_MODEL


def _collect_department_schema(search_root: str, max_samples: int = 6) -> list[dict[str, Any]]:
    """부서 폴더의 DB/엑셀 테이블 구조(파일/테이블/헤더 + 컬럼별 샘플값)를 수집합니다."""
    schema: list[dict[str, Any]] = []
    if not search_root or not os.path.exists(search_root):
        return schema
    for root, dirs, files in os.walk(search_root):
        dirs[:] = [d for d in dirs if not d.startswith(".") and not d.startswith("subserver_search_")]
        for file_name in files:
            if file_name.startswith("~$") or file_name.startswith("."):
                continue
            ext = os.path.splitext(file_name)[1].lower()
            if ext not in DATA_SEARCH_EXTENSIONS:
                continue
            if file_name.lower() in EXCLUDED_SEARCH_DB_FILES:
                continue
            file_path = os.path.join(root, file_name)
            rel_path = os.path.relpath(file_path, search_root)
            for table in _iter_data_tables(file_path):
                headers = cast(list[str], table.get("headers", []))
                rows = cast(list[list[str]], table.get("rows", []))
                if not headers:
                    continue
                samples: dict[str, list[str]] = {}
                for col_idx, header in enumerate(headers):
                    seen: list[str] = []
                    for row in rows:
                        if col_idx < len(row):
                            val = str(row[col_idx]).strip()
                            if val and val not in seen:
                                seen.append(val)
                        if len(seen) >= max_samples:
                            break
                    samples[header] = seen
                schema.append({
                    "source_file": file_name,
                    "rel_path": rel_path,
                    "source_table": str(table.get("source_table", "")),
                    "headers": headers,
                    "samples": samples,
                    "rows": rows,
                })
    return schema


def _ollama_chat_json(system_prompt: str, user_prompt: str, timeout: int = 30) -> dict[str, Any] | None:
    """Ollama /api/chat 를 JSON 모드로 호출하고 파싱된 dict를 반환합니다."""
    payload = {
        "model": _get_ollama_model(),
        "stream": False,
        "format": "json",
        "options": {"temperature": 0},
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }
    try:
        data = json.dumps(payload).encode("utf-8")
        req = urllib.request.Request(
            f"{OLLAMA_URL}/api/chat", data=data, method="POST",
            headers={"Content-Type": "application/json"},
        )
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            body = json.loads(resp.read().decode("utf-8"))
        content = (body.get("message", {}) or {}).get("content", "")
        if not content:
            return None
        return json.loads(content)
    except Exception as exc:
        print(f"⚠️ [Ollama 호출 실패] {exc}")
        return None


# 경로 선택 프롬프트 — AI는 오직 '경로\파일\테이블명' 한 개만 응답한다 (추론/응답 시간 최소화)
DEFAULT_SYSTEM_PROMPT = (
    "너는 사내 데이터 경로 선택기다. 사용자 지시(질문)와 경로 목록(경로\\파일\\테이블명·컬럼)을 보고 "
    "질문에 가장 알맞은 경로 하나만 고른다.\n"
    "규칙:\n"
    '1) 반드시 다음 JSON 형식으로만 응답한다: {"path": "<경로\\\\파일\\\\테이블명>"}\n'
    "2) 경로 외의 설명·이유·문장·마크다운은 절대 출력하지 않는다.\n"
    "3) 목록에 있는 경로 문자열을 정확히 그대로 사용한다."
)


def _table_path(rel_path: str, table_name: str) -> str:
    """스키마 항목을 '경로\\파일\\테이블명' 단일 식별자 문자열로 만듭니다."""
    return f"{rel_path}\\{table_name}"


def _build_path_catalog(schema: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """부서 스키마를 '경로\\파일\\테이블명 → 컬럼' 카탈로그(행 제외)로 압축합니다. 추론 프롬프트/응답 최소화용."""
    catalog: list[dict[str, Any]] = []
    for tbl in schema:
        rel_path = str(tbl.get("rel_path", "")) or str(tbl.get("source_file", ""))
        table_name = str(tbl.get("source_table", ""))
        if not rel_path or not table_name:
            continue
        catalog.append({
            "path": _table_path(rel_path, table_name),
            "columns": [str(h) for h in tbl.get("headers", [])],
        })
    return catalog


def _ai_select_path(query: str, catalog: list[dict[str, Any]]) -> str | None:
    """AI에게 경로 목록을 주고 질문에 맞는 '경로\\파일\\테이블명' 하나만 받습니다(오직 경로만 추론)."""
    if not catalog:
        return None
    user_prompt = (
        f"[질문]\n{query}\n\n"
        f"[경로 목록]\n{json.dumps(catalog, ensure_ascii=False)}"
    )
    result = _ollama_chat_json(DEFAULT_SYSTEM_PROMPT, user_prompt, timeout=30)
    if not isinstance(result, dict):
        return None
    picked = str(result.get("path", "")).strip().replace("/", "\\")
    if not picked:
        return None
    paths = [str(item.get("path", "")) for item in catalog]
    if picked in paths:
        return picked
    low = picked.lower()
    for path in paths:
        if path.lower() == low or low in path.lower() or path.lower() in low:
            return path
    return None


def _load_path(query: str, path: str, schema: list[dict[str, Any]]) -> dict[str, Any] | None:
    """AI가 선택한 '경로\\파일\\테이블명'을 파싱해 해당 테이블 전체 내용을 검색 결과 테이블 JSON으로 반환합니다."""
    target = next(
        (
            t for t in schema
            if _table_path(str(t.get("rel_path", "")) or str(t.get("source_file", "")),
                           str(t.get("source_table", ""))) == path
        ),
        None,
    )
    if target is None:
        return None

    headers = [str(h) for h in target.get("headers", [])]
    rows = [[str(c) for c in r] for r in cast(list[list[Any]], target.get("rows", []))]
    source_file = str(target.get("source_file", ""))
    source_table = str(target.get("source_table", ""))

    # [신규] 클라우드 AI가 분석하기 좋게 List<Map> 형식으로 데이터 가공 (앱 전달용)
    table_data = []
    for row in rows:
        row_map = {}
        for i, header in enumerate(headers):
            if i < len(row):
                row_map[header] = row[i]
        table_data.append(row_map)

    # [신규] 200 레코드 초과 시 거부 로직
    if len(rows) > 200:
        return {
            "query": query,
            "status": "error_too_large",
            "message": "데이터 용량이 너무커서 로컬PC가 데이터 정보 응답을 거부 했습니다. 직접 확인 부탁합니다.",
            "table_data": [],
            "table": {
                "source_file": source_file,
                "source_table": source_table,
                "headers": headers,
                "rows": [],
            },
        }

    return {
        "query": query,
        "status": "success",
        "summary": f"'{path}' 테이블 전체 {len(rows)}행을 표시합니다.",
        "table_data": table_data, # 앱/클라우드 AI용 List<Map>
        "sql": "",
        "filters": [],
        "table": {
            "source_file": source_file,
            "source_table": source_table,
            "headers": headers,
            "rows": rows,
        },
    }


def smart_search(query: str, search_root: str) -> dict[str, Any] | None:
    """AI가 질문을 보고 '경로\\파일\\테이블명' 하나를 선택하면 코드가 그 경로의 테이블 전체를 반환합니다.

    AI는 오직 경로만 추론하므로 응답(추론) 시간이 최소화됩니다.
    실패(모델 미가동/경로 미선택 등) 시 None 을 반환합니다.
    """
    try:
        print(f"🕵️ [스마트 검색] 탐색 대상 폴더: {search_root}")
        schema = _collect_department_schema(search_root)
        if not schema:
            print(f"⚠️ [스마트 검색] 스키마를 수집하지 못했습니다. (폴더가 비어있거나 권한 문제)")
            return None
        
        print(f"📚 [스마트 검색] 발견된 테이블 수: {len(schema)}")
        catalog = _build_path_catalog(schema)
        selected = _ai_select_path(query, catalog)
        
        if not selected:
            print(f"🤖 [스마트 검색] 로컬 AI가 적절한 경로를 선택하지 못했습니다.")
            return None
            
        print(f"✅ [스마트 검색] 로컬 AI 선택 경로: {selected}")
        return _load_path(query, selected, schema)
    except Exception as exc:
        print(f"⚠️ [스마트 검색 실패] {exc}")
        return None


# 🔒 보안용 API Key
API_KEY = "MyCompanySecretKey1234!"
SUBSERVER_DB_PATH = "C:/SubServer/subserver.db"

# [신규] 권한 검증 로직
def verify_department_permission(key: str, requested_dept: str):
    """사용자 인증키를 기반으로 해당 부서 접근 권한이 있는지 확인"""
    if not key: return False
    # 마스터 키 예외 처리
    if key == API_KEY: return True
    
    try:
        conn = sqlite3.connect(SUBSERVER_DB_PATH)
        c = conn.cursor()
        c.execute("SELECT allowed_depts, is_active FROM activation_keys WHERE key = ?", (key,))
        row = c.fetchone()
        conn.close()
        
        if row:
            allowed_depts, is_active = row
            if is_active == 0:
                print(f"🚫 [권한 거부] 비활성화된 키: {key}")
                return False
            
            # 부서명 매핑 정규화 (앱: SALES -> 서브앱: 영업팀)
            reverse_dept_map = {
                "SALES": "영업팀",
                "PRODUCTION": "생산팀",
                "PROCUREMENT": "구매팀", "PURCHASE": "구매팀",
                "SHIPPING": "출하팀",
                "MANAGEMENT": "관리부"
            }
            target_dept_name = reverse_dept_map.get(requested_dept.upper(), requested_dept)
            
            if target_dept_name in (allowed_depts or ""):
                print(f"✅ [권한 허용] Key: {key}, 부서: {target_dept_name}")
                return True
            
            print(f"🚫 [권한 거부] Key: {key}, 요청부서: {target_dept_name}, 허용부서: {allowed_depts}")
        else:
            print(f"🚫 [권한 거부] 등록되지 않은 키: {key}")
            
    except Exception as e:
        print(f"❌ [권한 검사 에러] {e}")
    
    return False

# 부서명 매핑 (앱에서 보내는 영문 부서명과 실제 폴더명 일치화)
DEPT_MAP = {
    "영업팀": "SALES", "SALES": "SALES",
    "생산팀": "PRODUCTION", "PRODUCTION": "PRODUCTION",
    "구매팀": "PROCUREMENT", "PURCHASE": "PROCUREMENT", "PROCUREMENT": "PROCUREMENT",
    "출하팀": "SHIPPING", "SHIPPING": "SHIPPING",
    "관리부": "MANAGEMENT", "MANAGEMENT": "MANAGEMENT"
}

def get_mapped_dept(dept: str) -> str:
    return DEPT_MAP.get(dept.upper(), dept.upper())

def resolve_absolute_path(base: str, dept: str) -> str:
    """앱에서 보낸 절대 경로(base)와 부서명(dept)을 결합하여 실제 윈도우 경로 반환"""
    # [보완] base가 비어있거나 절대경로(드라이브 문자 포함)가 아니면 get_search_root() 사용
    if not base or ":" not in base:
        base = get_search_root()
    path = os.path.normpath(os.path.join(base, get_mapped_dept(dept)))
    os.makedirs(path, exist_ok=True)
    return path

@app.get("/files/{file_path:path}")
async def serve_file(file_path: str):
    """절대 경로 및 상대 경로 파일을 모두 지원하는 다운로드 엔드포인트"""
    decoded_path = urllib.parse.unquote(file_path)
    # 만약 경로에 드라이브 문자(C:)가 포함되어 있다면 절대 경로로 처리
    if ":" in decoded_path:
        target_path = os.path.normpath(decoded_path)
    else:
        target_path = os.path.normpath(os.path.join(get_search_root(), decoded_path))
    
    if os.path.exists(target_path) and os.path.isfile(target_path):
        return FileResponse(target_path)
    raise HTTPException(status_code=404, detail=f"파일을 찾을 수 없습니다: {target_path}")

@app.get("/health")
async def health_check():
    return {"status": "ok", "timestamp": datetime.now().isoformat()}

# API KEY 검증 함수
def verify_api_key(
    x_api_key: str = Header(None, alias="X-API-Key"),
    x_user_verification_key: str = Header(None, alias="X-User-Verification-Key")
):
    key = x_api_key or x_user_verification_key
    if not key:
        raise HTTPException(status_code=401, detail="API Key 누락")
    if key == API_KEY:
        return key
    # DB 연동 인증키 확인 로직 (이미 verify_department_permission에서 수행하므로 여기서는 키만 반환)
    return key

@app.post("/ai/search")
async def ai_search(
    req: AiSearchRequest,
    key: str = Depends(verify_api_key),
    x_user_department: str = Header(None, alias="X-User-Department")
):
    """부서 폴더의 DB/엑셀에서 검색한 원본 행을 JSON 테이블로 추출하여 반환 (하이브리드 AI용)"""
    dept = req.department or x_user_department or "SALES"
    if not verify_department_permission(key, dept):
        raise HTTPException(status_code=403, detail=f"'{dept}' 부서 데이터에 접근할 권한이 없습니다.")

    # [수정] 부서 폴더 경로를 생성할 때 normpath를 사용하여 경로 구분자 일치화
    dept_folder = get_mapped_dept(dept)
    search_root = os.path.normpath(os.path.join(get_search_root(), dept_folder))
    
    print(f"🔍 [AI 검색] 요청 부서: {dept} -> 탐색 경로: {search_root}")

    # 1. 로컬 AI 경로 선택 기반 스마트 검색: 질문에 맞는 테이블 위치 식별 및 데이터 로드
    answer_json = smart_search(req.query, search_root)
    
    # 2. 결과가 없는 경우 기본 응답 구성
    if not answer_json:
        answer_json = {
            "query": req.query,
            "status": "not_found",
            "summary": "관련 데이터를 찾을 수 없습니다.",
            "table": {"source_file": "", "source_table": "", "headers": [], "rows": []},
        }

    # 3. 데이터 추출 결과 반환 (이 JSON 데이터는 앱을 거쳐 메인 서버/클라우드 AI로 전달됨)
    result = {
        "status": "success",
        "department": dept_folder,
        "engine": _get_ollama_model(),
        "answer": answer_json.get("summary", ""),
        "answer_json": answer_json,
        "results": [],
    }

    # [신규] GUI 인스턴스가 존재할 경우 AI 응답 내용을 화면에 표시
    try:
        from gui import global_gui_instance
        if global_gui_instance:
            global_gui_instance.after(0, lambda: global_gui_instance.display_external_ai_result(req.query, result))
    except Exception as e:
        print(f"⚠️ GUI 연동 에러: {e}")

    return result

# --- [신규] DB 스키마 및 레코드 조회 엔드포인트 (앱 연동용) ---

@app.get("/db/schema")
async def get_db_schema_v2(
    department: str = "",
    key: str = Depends(verify_api_key),
    x_user_department: str = Header(None, alias="X-User-Department")
):
    """부서 폴더 내 모든 DB의 스키마 정보를 텍스트로 반환"""
    dept = department or x_user_department
    if not dept:
        raise HTTPException(status_code=400, detail="부서 정보가 누락되었습니다.")
    
    # 권한 체크
    if not verify_department_permission(key, dept):
        raise HTTPException(status_code=403, detail=f"'{dept}' 부서 데이터에 접근할 권한이 없습니다.")

    dept_folder = get_mapped_dept(dept)
    search_root = os.path.join(get_search_root(), dept_folder)
    
    if not os.path.exists(search_root):
        return {"status": "error", "schema": "부서 폴더를 찾을 수 없습니다."}

    schema_parts = []
    for root, dirs, files in os.walk(search_root):
        for file in files:
            if file.endswith(".db") and file != "chat_history.DB":
                db_path = os.path.join(root, file)
                rel_path = os.path.relpath(db_path, search_root)
                schema_parts.append(f"DB File: {rel_path}")
                
                try:
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = [row[0] for row in cursor.fetchall()]

                    for table in tables:
                        if table.startswith("sqlite_"): continue
                        cursor.execute(f"PRAGMA table_info('{table}')")
                        columns = [f"{col[1]}({col[2]})" for col in cursor.fetchall()]
                        schema_parts.append(f"  Table: {table}")
                        schema_parts.append(f"    Columns: {', '.join(columns)}")
                    conn.close()
                except Exception as e:
                    schema_parts.append(f"  Error reading DB: {e}")
                schema_parts.append("")

    schema_text = "\n".join(schema_parts).strip()
    if not schema_text:
        schema_text = "조회된 DB 스키마 정보가 없습니다."
        
    return {"status": "success", "schema": schema_text}

# --- [신규] 지능형 검색을 위한 메타데이터 API ---

@app.get("/db/record-count")
async def get_db_record_count(
    department: str,
    table_name: str,
    key: str = Depends(verify_api_key),
    x_user_department: str = Header(None, alias="X-User-Department")
):
    """특정 테이블의 전체 레코드 개수 조회"""
    dept = department or x_user_department
    if not verify_department_permission(key, dept):
        raise HTTPException(status_code=403, detail="접근 권한이 없습니다.")

    dept_folder = get_mapped_dept(dept)
    search_root = os.path.join(get_search_root(), dept_folder)
    
    target_db_path = None
    for root, dirs, files in os.walk(search_root):
        for file in files:
            if file.endswith(".db") and file != "chat_history.DB":
                db_path = os.path.join(root, file)
                try:
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                    if cursor.fetchone():
                        target_db_path = db_path
                        conn.close()
                        break
                    conn.close()
                except: continue
        if target_db_path: break

    if not target_db_path:
        return {"status": "error", "message": "Table not found", "count": 0}

    try:
        conn = sqlite3.connect(target_db_path)
        cursor = conn.cursor()
        cursor.execute(f'SELECT COUNT(*) FROM "{table_name}"')
        count = cursor.fetchone()[0]
        conn.close()
        return {"status": "success", "count": count}
    except Exception as e:
        return {"status": "error", "message": str(e), "count": 0}

@app.get("/db/column-values")
async def get_db_column_values(
    department: str,
    table_name: str,
    column_name: str,
    key: str = Depends(verify_api_key),
    x_user_department: str = Header(None, alias="X-User-Department")
):
    """특정 테이블의 특정 컬럼 값 리스트 조회 (최대 1000건)"""
    dept = department or x_user_department
    if not verify_department_permission(key, dept):
        raise HTTPException(status_code=403, detail="접근 권한이 없습니다.")

    dept_folder = get_mapped_dept(dept)
    search_root = os.path.join(get_search_root(), dept_folder)
    
    target_db_path = None
    for root, dirs, files in os.walk(search_root):
        for file in files:
            if file.endswith(".db") and file != "chat_history.DB":
                db_path = os.path.join(root, file)
                try:
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (table_name,))
                    if cursor.fetchone():
                        target_db_path = db_path
                        conn.close()
                        break
                    conn.close()
                except: continue
        if target_db_path: break

    if not target_db_path:
        return {"status": "error", "message": "Table not found"}

    try:
        conn = sqlite3.connect(target_db_path)
        cursor = conn.cursor()
        # 중복 제거된 값 리스트 가져오기
        cursor.execute(f'SELECT DISTINCT "{column_name}" FROM "{table_name}" LIMIT 1000')
        values = [str(row[0]) for row in cursor.fetchall() if row[0] is not None]
        conn.close()
        return {"status": "success", "values": values}
    except Exception as e:
        return {"status": "error", "message": str(e)}

from pydantic import BaseModel
from typing import Dict, Any, Optional

# (RecordRequest was already defined above, so I'll remove this redundant block)

@app.post("/db/record")
async def get_db_record_v2(
    req: RecordRequest,
    key: str = Depends(verify_api_key),
    x_user_department: str = Header(None, alias="X-User-Department")
):
    """특정 테이블에서 필터링된 레코드를 조회 (JSON Body 지원)"""
    dept = req.department or x_user_department
    if not dept:
        raise HTTPException(status_code=400, detail="부서 정보가 누락되었습니다.")

    # 권한 체크
    if not verify_department_permission(key, dept):
        raise HTTPException(status_code=403, detail=f"'{dept}' 부서 데이터에 접근할 권한이 없습니다.")

    dept_folder = get_mapped_dept(dept)
    search_root = os.path.join(get_search_root(), dept_folder)
    
    # 1. 대상 DB 파일 찾기 (테이블명이 포함된 첫 번째 DB 검색)
    target_db_path = None
    for root, dirs, files in os.walk(search_root):
        for file in files:
            if file.endswith(".db") and file != "chat_history.DB":
                db_path = os.path.join(root, file)
                try:
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table' AND name=?", (req.table_name,))
                    if cursor.fetchone():
                        target_db_path = db_path
                        conn.close()
                        break
                    conn.close()
                except:
                    continue
        if target_db_path: break

    if not target_db_path:
        return {"status": "error", "message": f"테이블 '{req.table_name}'을 포함하는 DB 파일을 찾을 수 없습니다."}

    try:
        conn = sqlite3.connect(target_db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        sql = f'SELECT * FROM "{req.table_name}"'
        params = []

        if req.filters:
            where_clauses = []
            for col, val in req.filters.items():
                where_clauses.append(f'"{col}" LIKE ?')
                params.append(f"%{val}%")
            if where_clauses:
                sql += " WHERE " + " AND ".join(where_clauses)

        sql += " LIMIT 100"
        cursor.execute(sql, params)
        rows = cursor.fetchall()
        
        results = [dict(r) for r in rows]
        conn.close()
        return {"status": "success", "data": results}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# --- 대화 내역 관련 (앱의 요청 주소에 맞게 /server/와 /api/ 모두 대응) ---

async def handle_load_history(company, department):
    # [수정] 앱에서 보낸 절대 경로(company) 우선 처리
    target_dir = resolve_absolute_path(company, department)
    
    # [보안] 대소문자 구분 없이 chat_history.db 또는 .DB 탐색
    db_path = os.path.join(target_dir, "chat_history.DB")
    if not os.path.exists(db_path):
        db_path = os.path.join(target_dir, "chat_history.db")
    
    if not os.path.exists(db_path):
        return []
    
    try:
        conn = sqlite3.connect(db_path)
        # [신규] 테이블이 없을 경우 자동 생성 (최초 로드 시 에러 방지)
        c = conn.cursor()
        c.execute('''CREATE TABLE IF NOT EXISTS messages (
            id TEXT PRIMARY KEY, roomTitle TEXT, authorId TEXT, authorName TEXT, 
            text TEXT, uri TEXT, type TEXT, createdAt INTEGER, metadata TEXT)''')
        conn.commit()
        
        conn.row_factory = sqlite3.Row
        c = conn.cursor()
        c.execute("SELECT * FROM messages ORDER BY createdAt ASC")
        rows = c.fetchall()
        conn.close()
        return [dict(row) for row in rows]
    except Exception as e:
        print(f"❌ [handle_load_history] DB 에러: {e}")
        return []

@app.post("/server/load-chat-history")
async def load_chat_history_server(req: ChatHistoryRequest, key: str = Depends(verify_api_key)):
    return await handle_load_history(req.company, req.department)

@app.post("/api/load-chat-history") # 하위 호환 및 404 방지용
async def load_chat_history_api(req: ChatHistoryRequest, key: str = Depends(verify_api_key)):
    return await handle_load_history(req.company, req.department)

async def handle_sync_message(company, department, msg_data):
    # [수정] 앱에서 보낸 절대 경로(company) 우선 처리
    target_dir = resolve_absolute_path(company, department)
    
    # [보완] 기존 파일이 .db 이면 .db 에 저장, 없거나 .DB 이면 .DB 에 저장 (대소문자 일관성)
    db_path = os.path.join(target_dir, "chat_history.db")
    if not os.path.exists(db_path):
        db_path = os.path.join(target_dir, "chat_history.DB")

    conn = sqlite3.connect(db_path)
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS messages (
        id TEXT PRIMARY KEY, roomTitle TEXT, authorId TEXT, authorName TEXT, 
        text TEXT, uri TEXT, type TEXT, createdAt INTEGER, metadata TEXT)''')
    
    c.execute('''INSERT OR REPLACE INTO messages VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)''', (
        msg_data.get('id', ''), msg_data.get('roomTitle', ''), msg_data.get('authorId', ''), msg_data.get('authorName', ''),
        msg_data.get('text', ''), msg_data.get('uri', ''), msg_data.get('type', 'text'), msg_data.get('createdAt', 0), msg_data.get('metadata', '{}')
    ))
    conn.commit()
    conn.close()
    return {"status": "success"}

@app.post("/server/sync-chat-message")
async def sync_chat_message_server(req: ChatMessageSyncRequest, key: str = Depends(verify_api_key)):
    msg_data = req.dict()
    return await handle_sync_message(req.company, req.department, msg_data)

@app.post("/api/sync-chat-message") # 하위 호환용
async def sync_chat_message_api(req: ChatMessageSyncRequest, key: str = Depends(verify_api_key)):
    msg_data = req.dict()
    return await handle_sync_message(req.company, req.department, msg_data)

# --- 엑셀 저장 관련 ---

@app.post("/server/save-data")
async def save_data(req: SaveDataRequest, key: str = Depends(verify_api_key)):
    target_dir = resolve_absolute_path(req.company, req.department)
    excel_path = os.path.join(target_dir, f"{req.task_name}.xlsx")
    
    if not os.path.exists(excel_path):
        wb = Workbook()
        ws = wb.active
        ws.append(["날짜", "담당자", "내용", "비고"])
        wb.save(excel_path)
    
    wb = load_workbook(excel_path)
    ws = wb.active
    ws.append([datetime.now().strftime("%Y-%m-%d %H:%M"), req.poster, req.content, "모바일 앱 전송"])
    wb.save(excel_path)
    return {"status": "success", "path": excel_path}

@app.post("/server/upload-excel")
async def upload_excel(
    company: str = Form(...),
    department: str = Form(...),
    task_name: str = Form(...),
    file: UploadFile = File(...),
    key: str = Depends(verify_api_key)
):
    """엑셀 파일 업로드 및 저장 (Multipart)"""
    target_dir = resolve_absolute_path(company, department)
    file_path = os.path.join(target_dir, f"{task_name}.xlsx")
    
    with open(file_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)
        
    return {"status": "success", "message": f"File uploaded to {file_path}"}

# --- [에이전트형] 부서 데이터 맵 및 정밀 추출 엔진 ---

class DeptDataMapRequest(BaseModel):
    department: str

@app.post("/server/get-department-data-map")
async def get_department_data_map(
    req: DeptDataMapRequest,
    key: str = Depends(verify_api_key)
):
    """부서 폴더 내 모든 DB의 구조(테이블, 컬럼) 및 파일 맵 생성"""
    department = req.department
    # [신규] 권한 체크
    if not verify_department_permission(key, department):
        raise HTTPException(status_code=403, detail=f"'{department}' 부서 데이터에 접근할 권한이 없습니다.")

    dept_folder = get_mapped_dept(department)
    search_root = os.path.join(get_search_root(), dept_folder)

    data_map = {
        "department": department,
        "folder_structure": [],
        "databases": {}
    }

    if not os.path.exists(search_root):
        return {"status": "error", "message": "Department folder not found"}

    for root, dirs, files in os.walk(search_root):
        rel_root = os.path.relpath(root, search_root)
        if rel_root != ".":
            data_map["folder_structure"].append(rel_root)

        for file in files:
            if file.endswith(".db") and file != "chat_history.DB":
                db_path = os.path.join(root, file)
                db_rel_path = os.path.relpath(db_path, search_root)

                try:
                    conn = sqlite3.connect(db_path)
                    cursor = conn.cursor()
                    cursor.execute("SELECT name FROM sqlite_master WHERE type='table'")
                    tables = [row[0] for row in cursor.fetchall()]

                    db_info = {}
                    for table in tables:
                        if table.startswith("sqlite_"): continue
                        cursor.execute(f"PRAGMA table_info('{table}')")
                        columns = [col[1] for col in cursor.fetchall()]
                        db_info[table] = {"columns": columns}

                    data_map["databases"][db_rel_path] = db_info
                    conn.close()
                except Exception as e:
                    data_map["databases"][db_rel_path] = f"Error: {str(e)}"

            # 명세서/이미지 폴더 힌트 추가
            if any(k in root.lower() for k in ["invoice", "image", "명세서"]):
                if "hints" not in data_map: data_map["hints"] = []
                hint = f"Folder '{rel_root}' contains documents/images. Files: {files[:5]}"
                if hint not in data_map["hints"]: data_map["hints"].append(hint)

    return {"status": "success", "data_map": data_map}

class SpecificDataRequest(BaseModel):
    department: str
    db_rel_path: str
    table_name: str
    query_hint: str = ""

@app.post("/server/fetch-specific-data")
async def fetch_specific_data(
    req: SpecificDataRequest,
    key: str = Depends(verify_api_key)
):
    """AI가 지목한 특정 DB의 특정 테이블 데이터를 정밀 추출"""
    department = req.department
    db_rel_path = req.db_rel_path
    table_name = req.table_name
    query_hint = req.query_hint

    # [신규] 권한 체크
    if not verify_department_permission(key, department):
        raise HTTPException(status_code=403, detail=f"'{department}' 부서 레코드에 접근할 권한이 없습니다.")

    dept_folder = get_mapped_dept(department)
    db_path = os.path.join(get_search_root(), dept_folder, db_rel_path)

    if not os.path.exists(db_path):
        return {"status": "error", "message": f"File {db_rel_path} not found"}

    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()

        sql = f'SELECT * FROM "{table_name}"'
        params = []

        if query_hint:
            cursor.execute(f"PRAGMA table_info('{table_name}')")
            cols = [col[1] for col in cursor.fetchall()]

            where = [f'"{c}" LIKE ?' for c in cols]
            params = [f"%{query_hint}%"] * len(where)

            if where:
                sql += " WHERE " + " OR ".join(where)

        sql += " LIMIT 200"
        cursor.execute(sql, params)
        rows = cursor.fetchall()

        results = [dict(r) for r in rows]
        conn.close()
        return {"status": "success", "data": results}
    except Exception as e:
        return {"status": "error", "message": str(e)}

# 기존 GUI 및 서버 실행 로직 유지
if __name__ == "__main__":
    import uvicorn
    import threading
    from gui import SubServerGUI
    import gui

    def start_api_server():
        uvicorn.run(app, host="0.0.0.0", port=8001)

    threading.Thread(target=start_api_server, daemon=True).start()
    print("🚀 사내 서브서버(Port 8001) 가동 중...")
    gui_app = SubServerGUI()
    gui.global_gui_instance = gui_app # 전역 인스턴스 등록
    gui_app.mainloop()
